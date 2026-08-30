########## Attendance Tracker ########


from email.message import EmailMessage
import smtplib
import openpyxl


wb = openpyxl.Workbook()
ws = wb.active

ws.column_dimensions["B"].width = 25

ws.append(["rollno", "mailid", "CI", "python", "DM"])
ws.append([1, "aaaa@gmail.com", 0, 0, 0])
ws.append([2, "bbbb@gmail.com", 0, 0, 0])
ws.append([3, "cccc@gmail.com", 0, 0, 0])

wb.save("attendance.xlsx")

def save():
    wb.save("attendance.xlsx")
    print("fichier sauvegardé!")

def send_mail(msg, to, subject_no):
    mail = EmailMessage()
    mail["FROM"] = "xxxxx@gmail.com"
    mail["TO"] = to
    mail["SUBJECT"] = f"Liste de presence en {check(subject_no)}"

    mail.set_content(msg)

    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login("xxxxx@gmail.com", "password")
        smtp.send_message(msg)

def check(i):
    global sub

    if i == 1:
        sub = "CI"
    elif i == 2:
        sub = "python"
    elif i == 3:
        sub = "DM"
    return sub

resp = 1
msgs = ["Deja 2 absences.Plus qu'une et pas d'examen!",
        "3 absences. Pas d'examen!"]

while resp == 1:
    print("CI-->1\npython-->2\nDM-->3")
    sub = int(input("entrer la matiere:"))

    for n in range(1, 4):
        c = int(input(f"l'etudiant {n} est il absent? oui-->1, non-->0"))
        if c == 1:
            ws.cell(row=n+1, column=sub+2).value += 1
            save()
        if ws.cell(row=n+1, column=sub+2).value == 2:
            send_mail(msgs[0], ws.cell(row=n+1, column=2).value, sub)
            print("mail envoyé!")
        elif ws.cell(row=n+1, column=sub+2).value > 2:
            send_mail(msgs[1], ws.cell(row=n+1, column=2).value, sub)
            send_mail(f"Pas d'examen pour l'etudiant {n}",
                      "staff@gmail.com", sub)
            print("mails envoyés!")

    resp = int(input("Autre matiere? Oui-->1, Non-->0"))





